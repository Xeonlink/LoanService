from datetime import datetime
from openpyxl import Workbook
import sqlite3
import sys
import os
import re


def swapkv(d: dict):
    return {v: k for k, v in d.items()}


def readlines(*path):
    with open(os.path.join(*path), "r") as f:
        return "".join(f.readlines())


def format_loan_duration(start: datetime, end: datetime):
    if start.year == end.year:
        if start.month == end.month:
            if start.day == end.day:
                return start.strftime("%Y.%m.%d")
            else:
                return f"{start.strftime('%Y.%m.%d')} ~ {end.strftime('%m.%d')}"
        else:
            return f"{start.strftime('%Y.%m.%d')} ~ {end.strftime('%m.%d')}"
    else:
        return f"{start.strftime('%Y.%m.%d')} ~ {end.strftime('%Y.%m.%d')}"


def resource_path(relative_path: str) -> str:
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        base_path = sys._MEIPASS  # type: ignore
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def sanitize_sheet_name(name):
    # Excel 워크시트 이름으로 사용 가능한 문자열로 변환
    name = re.sub(r"[\\/*?:\[\]]", "_", name)
    return name[:31]  # 최대 길이 31자 제한


def sqlite_to_excel(db_path: str, excel_path: str):
    """
    SQLite 데이터베이스의 각 테이블을 Excel 워크시트로 내보냅니다.

    :db_path: SQLite 데이터베이스 파일 경로 (str)
    :excel_path: 생성할 Excel 파일 경로 (str)
    """
    with sqlite3.connect(db_path) as conn:
        cursor = conn.cursor()

        # 데이터베이스 내의 모든 테이블 이름을 가져옵니다
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()

        # 새로운 Excel 워크북을 생성합니다
        workbook = Workbook()

        # 기본 시트를 제거합니다 (시트 이름이 'Sheet'인 경우)
        if 'Sheet' in workbook.sheetnames:
            default_sheet = workbook['Sheet']
            workbook.remove(default_sheet)

        for table_name in tables:
            # 테이블 이름을 추출합니다
            table_name = table_name[0]

            # 테이블 이름을 안전하게 처리합니다
            safe_table_name = f"'{table_name.replace('\'', '\'\'')}'"

            # 테이블의 모든 데이터를 가져옵니다
            cursor.execute(f"SELECT * FROM {safe_table_name}")
            rows = cursor.fetchall()

            # 테이블의 열 이름을 가져옵니다
            cursor.execute(f"PRAGMA table_info({safe_table_name})")
            columns = [info[1] for info in cursor.fetchall()]

            # 워크시트 이름을 정제합니다 (특수 문자 제거 및 길이 제한)
            sheet_name = re.sub(r'[\\/*?:\[\]]', '_', table_name)[:31]

            # 새로운 워크시트를 추가합니다
            worksheet = workbook.create_sheet(title=sheet_name)

            # 열 이름을 첫 번째 행에 작성합니다
            worksheet.append(columns)

            # 데이터를 워크시트에 작성합니다
            for row in rows:
                worksheet.append(row)

        # Excel 파일을 저장합니다
        workbook.save(excel_path)
